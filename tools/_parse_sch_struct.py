"""Parse C-struct rows from official PNE SCH xlsx."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

SCH_PATH = Path(r"c:\(PNE) LG sch 스케줄 파일 구조_20250211_ 1.xlsx")

SIZES = {
    "BYTE": 1,
    "BOOL": 4,
    "char": 1,
    "WORD": 2,
    "UINT": 4,
    "ULONG": 4,
    "int": 4,
    "LONG": 4,
    "float": 4,
    "double": 8,
}


def field_size(dtype: str) -> int:
    dtype = str(dtype).strip()
    base = dtype.split("[")[0].strip()
    if "[" in dtype:
        m = re.search(r"\[(\d+)\]", dtype)
        n = int(m.group(1)) if m else 1
        return SIZES.get(base, 1) * n
    return SIZES.get(base, SIZES.get(dtype, 4))


def parse_step_struct(ws, max_rows: int = 250) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))[:max_rows]
    fields: list[dict] = []
    offset = 0
    in_step = False
    for i, row in enumerate(rows):
        col0 = row[0] if row else None
        section = row[1] if row and len(row) > 1 else None
        dtype = row[3] if row and len(row) > 3 else None
        name = row[4] if row and len(row) > 4 else None
        if col0 and "STEP" in str(col0).upper():
            in_step = True
            offset = 0
            continue
        if section == "FILE_STEP_CONDITION":
            in_step = True
            offset = 0
            continue
        if not in_step or not dtype or not name:
            continue
        sz = field_size(dtype)
        fields.append(
            {
                "offset": offset,
                "size": sz,
                "dtype": str(dtype).strip(),
                "name": str(name).strip().rstrip(";"),
            }
        )
        offset += sz
    return fields


def main() -> None:
    wb = openpyxl.load_workbook(SCH_PATH, read_only=True, data_only=True)
    hot = {12, 16, 20, 24, 28, 32, 36, 48, 52, 92, 392, 428, 496, 512, 513, 564}
    for sn in wb.sheetnames:
        fields = parse_step_struct(wb[sn])
        if not fields:
            print(f"=== {sn}: no step struct parsed")
            continue
        total = fields[-1]["offset"] + fields[-1]["size"]
        print(f"=== {sn}: {len(fields)} fields, naive struct size = {total} bytes")
        for f in fields:
            mark = " <<" if f["offset"] in hot else ""
            print(f"  +{f['offset']:4d}  {f['size']:2d}  {f['dtype']:18s}  {f['name']}{mark}")
        print()
    wb.close()


if __name__ == "__main__":
    main()
