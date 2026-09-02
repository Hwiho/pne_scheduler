"""One-off analysis of three PNE Excel structure files."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

FILES = {
    "sch_official": Path(r"c:\(PNE) LG sch 스케줄 파일 구조_20250211_ 1.xlsx"),
    "data_pne": Path(r"c:\pne_data_structure.xlsx"),
    "data_structure_pne": Path(r"c:\data_structure_pne.xlsx"),
}
OUT = Path(__file__).resolve().parents[1] / "example" / "reports" / "xlsx_three_file_analysis.json"


def read_sheet_matrix(ws, max_rows: int = 200, max_cols: int = 20) -> list[list]:
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append(list(row[:max_cols]))
    return rows


def find_header_row(rows: list[list]) -> tuple[int, list]:
    for i, row in enumerate(rows[:30]):
        texts = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(texts)
        if any(k in joined for k in ("offset", "name", "type", "size", "field", "byte", "variable")):
            return i, row
    return 0, rows[0] if rows else []


def extract_fields_from_sch_sheet(rows: list[list]) -> list[dict]:
    hi, _header = find_header_row(rows)
    fields: list[dict] = []
    for row in rows[hi + 1 :]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        vals = [row[j] if j < len(row) else None for j in range(min(12, len(row)))]
        nonempty = [v for v in vals if v is not None and str(v).strip()]
        if len(nonempty) <= 1:
            sec = str(nonempty[0]) if nonempty else None
            if sec and len(sec) > 3:
                fields.append({"kind": "section", "name": sec})
            continue

        offset = None
        off = vals[0]
        if isinstance(off, (int, float)) and off < 10000:
            offset = int(off)
        elif isinstance(off, str):
            s = off.strip().lower()
            if s.startswith("0x"):
                offset = int(s, 16)
            elif s.replace(".", "").isdigit():
                offset = int(float(s))

        name_cand = None
        for v in vals[1:6]:
            if v is None:
                continue
            vs = str(v).strip()
            if vs and not vs.replace(".", "").replace("-", "").isdigit():
                name_cand = vs
                break

        dtype_cand = None
        for v in vals[2:8]:
            if v is None:
                continue
            vs = str(v).strip().upper()
            if vs in ("FLOAT", "INT", "ULONG", "BOOL", "BYTE", "WORD", "DWORD", "SHORT", "UINT", "DOUBLE", "CHAR"):
                dtype_cand = vs
                break

        if offset is not None or name_cand:
            fields.append(
                {
                    "kind": "field",
                    "offset": offset,
                    "name": name_cand,
                    "dtype": dtype_cand,
                    "raw": [str(v) if v is not None else "" for v in vals[:8]],
                }
            )
    return fields


def analyze_data_structure_sheet(ws, max_rows: int = 120) -> dict:
    rows = read_sheet_matrix(ws, max_rows=max_rows, max_cols=15)
    hi, header = find_header_row(rows)
    preview = [[str(c)[:40] if c is not None else "" for c in row[:10]] for row in rows[:15]]
    var_names: list[str] = []
    for row in rows[hi + 1 :]:
        if not row:
            continue
        name = row[1] if len(row) > 1 else row[0]
        if name and str(name).strip() and str(name).strip().lower() not in ("name", "variable", "항목"):
            var_names.append(str(name).strip())
    return {
        "header_row": hi,
        "header": [str(c) if c is not None else "" for c in (header or [])[:12]],
        "preview": preview,
        "sample_vars": var_names[:30],
        "var_count_approx": len(var_names),
    }


def main() -> None:
    report: dict = {}

    sch_wb = openpyxl.load_workbook(FILES["sch_official"], read_only=True, data_only=True)
    sch_report: dict = {"file": FILES["sch_official"].name, "sheets": {}}
    for sn in sch_wb.sheetnames:
        ws = sch_wb[sn]
        rows = read_sheet_matrix(ws, max_rows=500, max_cols=16)
        fields = extract_fields_from_sch_sheet(rows)
        sections = [f["name"] for f in fields if f.get("kind") == "section"]
        flds = [f for f in fields if f.get("kind") == "field" and f.get("offset") is not None]
        hot = {off: [] for off in [12, 16, 20, 24, 28, 32, 48, 52, 496, 512, 564]}
        for f in flds:
            o = f["offset"]
            if o in hot:
                hot[o].append(f)
        sch_report["sheets"][sn] = {
            "sections": sections[:20],
            "field_count": len(flds),
            "hot_offsets": {str(k): v for k, v in hot.items() if v},
            "first_fields": flds[:15],
            "preview_row0": [str(c)[:40] if c is not None else "" for c in rows[0][:8]] if rows else [],
        }
    sch_wb.close()
    report["sch_official"] = sch_report

    for key in ("data_pne", "data_structure_pne"):
        wb = openpyxl.load_workbook(FILES[key], read_only=True, data_only=True)
        dr: dict = {"file": FILES[key].name, "sheets": {}}
        for sn in wb.sheetnames:
            dr["sheets"][sn] = analyze_data_structure_sheet(wb[sn])
        wb.close()
        report[key] = dr

    s1 = set(report["data_pne"]["sheets"].keys())
    s2 = set(report["data_structure_pne"]["sheets"].keys())
    report["data_file_comparison"] = {
        "only_in_pne_data_structure": sorted(s1 - s2),
        "only_in_data_structure_pne": sorted(s2 - s1),
        "shared": sorted(s1 & s2),
    }

    ensol = {
        12: "volt_or_vlim_mV",
        16: "current_mA",
        20: "time_or_rest_s",
        28: "voltage_cutoff_mV",
        32: "cv_cutoff_mA",
    }
    sheet_10003 = sch_report["sheets"].get("0x00010003", {})
    report["ensol_vs_official_10003"] = [
        {
            "offset": off,
            "ensol": ensol_name,
            "official_xlsx": [e.get("name") for e in sheet_10003.get("hot_offsets", {}).get(str(off), [])],
            "raw": sheet_10003.get("hot_offsets", {}).get(str(off), []),
        }
        for off, ensol_name in ensol.items()
    ]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUT}")
    for sn, info in sch_report["sheets"].items():
        print(f"  SCH {sn}: {info['field_count']} fields")


if __name__ == "__main__":
    main()
